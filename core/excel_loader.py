"""Load Excel workbooks and detect phone-number tables."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.message import FALLBACK_NAME
from core.phone import is_valid_phone, normalize_phone

PHONE_HEADER_PATTERNS = [
    r"mobiel\s*telefoon",
    r"telefoonnummer",
    r"telefoon",
    r"mobiel",
    r"gsm",
    r"phone",
    r"mobile",
    r"cell",
]

NAME_HEADER_PATTERNS = [
    r"volledige\s*naam",
    r"naam",
    r"name",
    r"student",
    r"voornaam",
]

SENT_HEADER_PATTERNS = [
    r"bericht\s*verzonden",
    r"uitnodiging\s*verzonden",
    r"verzonden",
    r"verstuurd",
    r"uitgenodigd",
    r"invited",
    r"sent",
]

_TRUTHY = {
    "true", "waar", "ja", "yes", "y", "1", "x", "✓", "✔", "v", "done", "ok",
    "verzonden", "verstuurd",
}


def parse_bool(value) -> bool:
    """Interpret a cell value (incl. Excel checkbox TRUE/FALSE) as a boolean."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if not text:
        return False
    return text in _TRUTHY


@dataclass
class ContactRow:
    phone_raw: str
    phone_normalized: str
    name: str
    row_index: int
    already_sent: bool = False


@dataclass
class SheetTable:
    sheet_name: str
    headers: list[str]
    header_row: int
    rows: list[dict[str, str]]
    row_numbers: list[int]


def _cell_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _score_header(header: str, patterns: list[str]) -> int:
    header_lower = header.lower()
    best = 0
    for pattern in patterns:
        if re.search(pattern, header_lower):
            best = max(best, len(pattern))
    return best


def detect_header_row(sheet: Worksheet, max_scan: int = 20) -> Optional[int]:
    """Find the first row that looks like a table header (several non-empty cells)."""
    best_row = None
    best_count = 0
    for row_idx in range(1, min(max_scan, sheet.max_row or 1) + 1):
        values = [_cell_value(sheet.cell(row=row_idx, column=col).value)
                  for col in range(1, (sheet.max_column or 1) + 1)]
        non_empty = [v for v in values if v]
        if len(non_empty) >= 2 and len(non_empty) > best_count:
            best_count = len(non_empty)
            best_row = row_idx
    return best_row


def load_sheet_table(path: str | Path, sheet_name: str) -> SheetTable:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
    sheet = wb[sheet_name]
    header_row = detect_header_row(sheet)
    if header_row is None:
        wb.close()
        raise ValueError("Could not detect a table header in this sheet.")

    headers: list[str] = []
    header_columns: list[int] = []
    for col in range(1, (sheet.max_column or 1) + 1):
        val = _cell_value(sheet.cell(row=header_row, column=col).value)
        if val:
            headers.append(val)
            header_columns.append(col)

    if not headers:
        wb.close()
        raise ValueError("No column headers found in the detected header row.")

    rows: list[dict[str, str]] = []
    row_numbers: list[int] = []
    for row_idx in range(header_row + 1, (sheet.max_row or header_row) + 1):
        row_data: dict[str, str] = {}
        has_data = False
        for header, col_idx in zip(headers, header_columns):
            val = _cell_value(sheet.cell(row=row_idx, column=col_idx).value)
            row_data[header] = val
            if val:
                has_data = True
        if has_data:
            rows.append(row_data)
            row_numbers.append(row_idx)

    wb.close()
    return SheetTable(
        sheet_name=sheet_name,
        headers=headers,
        header_row=header_row,
        rows=rows,
        row_numbers=row_numbers,
    )


def list_sheet_names(path: str | Path) -> list[str]:
    wb = load_workbook(path, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def guess_phone_column(headers: list[str]) -> Optional[str]:
    best_header = None
    best_score = 0
    for header in headers:
        score = _score_header(header, PHONE_HEADER_PATTERNS)
        if score > best_score:
            best_score = score
            best_header = header
    return best_header


def guess_name_column(headers: list[str]) -> Optional[str]:
    best_header = None
    best_score = 0
    for header in headers:
        if _score_header(header, PHONE_HEADER_PATTERNS) > 0:
            continue
        if _score_header(header, SENT_HEADER_PATTERNS) > 0:
            continue
        score = _score_header(header, NAME_HEADER_PATTERNS)
        if score > best_score:
            best_score = score
            best_header = header
    return best_header


def guess_sent_column(headers: list[str]) -> Optional[str]:
    best_header = None
    best_score = 0
    for header in headers:
        if _score_header(header, PHONE_HEADER_PATTERNS) > 0:
            continue
        score = _score_header(header, SENT_HEADER_PATTERNS)
        if score > best_score:
            best_score = score
            best_header = header
    return best_header


def extract_contacts(
    table: SheetTable,
    phone_column: str,
    name_column: Optional[str] = None,
    country_code: str = "+31",
    sent_column: Optional[str] = None,
) -> list[ContactRow]:
    contacts: list[ContactRow] = []
    seen_phones: set[str] = set()

    for row, row_number in zip(table.rows, table.row_numbers):
        raw = row.get(phone_column, "")
        if not raw:
            continue

        normalized = normalize_phone(raw, country_code)
        if not normalized or not is_valid_phone(normalized):
            continue

        if normalized in seen_phones:
            continue
        seen_phones.add(normalized)

        if name_column:
            name = row.get(name_column, "") or FALLBACK_NAME
        else:
            name = FALLBACK_NAME

        already_sent = parse_bool(row.get(sent_column, "")) if sent_column else False

        contacts.append(
            ContactRow(
                phone_raw=raw,
                phone_normalized=normalized,
                name=name,
                row_index=row_number,
                already_sent=already_sent,
            )
        )

    return contacts


def mark_rows_sent(
    path: str | Path,
    sheet_name: str,
    column_header: str,
    header_row: int,
    row_numbers: list[int],
    value: bool = True,
) -> int:
    """Write a boolean (e.g. checkbox TRUE) into the sent-column for given rows.

    Returns the number of rows updated. Raises on file/lock errors so the
    caller can inform the user (e.g. when the workbook is still open in Excel).
    """
    if not row_numbers:
        return 0

    wb = load_workbook(path)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
        sheet = wb[sheet_name]

        col_idx = None
        for col in range(1, (sheet.max_column or 1) + 1):
            if _cell_value(sheet.cell(row=header_row, column=col).value) == column_header:
                col_idx = col
                break
        if col_idx is None:
            raise ValueError(f"Column '{column_header}' not found in header row.")

        for row_number in row_numbers:
            sheet.cell(row=row_number, column=col_idx, value=bool(value))

        wb.save(path)
    finally:
        wb.close()

    return len(row_numbers)
